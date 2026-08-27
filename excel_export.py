"""Excel 导出模块（支持把原图缩略图嵌入表格）"""
import os
import re
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
try:
    from PIL import Image as PILImage
except ImportError:  # Template-only exports do not require Pillow.
    PILImage = None
from config import Config


# The inventory workbook is kept next to the application.  Its two sheets use
# different column layouts, so records must be merged into the existing rows
# instead of being exported to a newly-created generic sheet.
TEMPLATE_FILENAME = "YCSP电脑信息.xlsx"
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), TEMPLATE_FILENAME)

WORKSTATION_COLUMNS = {
    "machine_name": "B", "nic1_name": "F", "nic1_mac": "G",
    "disk_gb": "H", "memory_gb": "I", "sn_number": "J",
}
OFFICE_COLUMNS = {
    "machine_name": "B", "nic1_name": "F", "nic1_mac": "G",
    "nic2_name": "H", "nic2_mac": "I", "disk_gb": "J",
    "memory_gb": "K", "sn_number": "L",
}


COLUMNS = [
    ('机器名', 'machine_name', 18),
    ('SN号', 'sn_number', 18),
    ('网卡1', 'nic1_name', 40),
    ('网卡1 MAC', 'nic1_mac', 22),
    ('网卡2', 'nic2_name', 40),
    ('网卡2 MAC', 'nic2_mac', 22),
    ('硬盘(GB)', 'disk_gb', 12),
    ('内存(GB)', 'memory_gb', 12),
    ('上传时间', 'upload_time', 22),
    ('原图', '_image', 34),   # 特殊列：嵌入图片缩略图
]

# 缩略图缓存目录
THUMB_DIR = os.path.join(Config.DATA_FOLDER, 'image_cache', 'thumbs')
# 缩略图最大尺寸（像素）
THUMB_MAX_W, THUMB_MAX_H = 260, 130
# 数据行行高（磅），1 磅 ≈ 1.333 px，100 磅 ≈ 133px 可容纳缩略图
ROW_HEIGHT_PT = 100


def _make_thumb(image_path: str) -> str:
    """生成（或复用缓存的）缩略图，返回缩略图路径；失败返回 None"""
    if PILImage is None or not image_path or not os.path.exists(image_path):
        return None
    os.makedirs(THUMB_DIR, exist_ok=True)
    thumb_path = os.path.join(THUMB_DIR, os.path.basename(image_path))
    if os.path.exists(thumb_path):
        return thumb_path
    try:
        img = PILImage.open(image_path)
        if img.mode not in ('RGB', 'L'):
            img = img.convert('RGB')
        # 只缩小不放大
        img.thumbnail((THUMB_MAX_W, THUMB_MAX_H), PILImage.LANCZOS)
        img.save(thumb_path, 'JPEG', quality=85)
        return thumb_path
    except Exception:
        return None


def export_to_excel(records: list, output_path: str = None,
                    embed_images: bool = True,
                    sort_by_machine_name: bool = True) -> str:
    """把识别记录导出为 Excel 文件，返回文件路径

    embed_images=True 时，每行末尾嵌入对应上传图片的缩略图。
    """
    if output_path is None:
        output_path = Config.EXCEL_PATH
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "OCR 识别结果"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写表头
    for col_idx, (title, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 写数据
    ordered_records = sorted(records or [], key=_machine_sort_key) if sort_by_machine_name else (records or [])
    for row_idx, rec in enumerate(ordered_records, start=2):
        # 行高设为容纳缩略图
        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

        for col_idx, (_title, key, _w) in enumerate(COLUMNS, start=1):
            # 特殊列：嵌入图片缩略图
            if key == '_image':
                if embed_images:
                    thumb = _make_thumb(rec.get('image_path', ''))
                    if thumb:
                        try:
                            img = XLImage(thumb)
                            img.anchor = f"{get_column_letter(col_idx)}{row_idx}"
                            ws.add_image(img)
                        except Exception:
                            pass  # 单张图片失败不影响整体导出
                continue

            # 与前端表格保持一致：空字段显示为短横线。
            value = rec.get(key) or '-'
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    # 冻结表头
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    return output_path


def _record_type(record: dict) -> str:
    """Return the template sheet type for an OCR record.

    Machine names are the reliable discriminator in the supplied inventory:
    ``YCSPGW`` denotes a workstation and ``YCSPBG`` an office computer.
    An explicit ``computer_type`` value is accepted for integrations that do
    not use those prefixes.
    """
    explicit = str(record.get("computer_type", "")).strip().lower()
    if explicit in {"workstation", "工位电脑", "gw"}:
        return "workstation"
    if explicit in {"office", "办公电脑", "bg"}:
        return "office"
    name = str(record.get("machine_name", "")).strip().upper()
    if re.match(r"YCSPGW", name):
        return "workstation"
    if re.match(r"YCSPBG", name):
        return "office"
    return ""


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    """Copy the visible row formatting when a new machine is appended."""
    if source_row < 1:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def export_to_legacy_template(records: list, template_path: str = TEMPLATE_PATH,
                              output_path: str = None,
                              sort_by_machine_name: bool = True) -> str:
    """Merge OCR records into the supplied two-sheet inventory template.

    Existing rows are matched by machine name and only fields present in a
    record are written.  This preserves manually-entered department/person
    data and all formatting.  Unknown machines are appended to the matching
    sheet with the next serial number.
    """
    if output_path is None:
        output_path = os.path.join(Config.DATA_FOLDER, "YCSP电脑信息_已填写.xlsx")
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"Excel template not found: {template_path}")
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = load_workbook(template_path)
    sheet_specs = {
        "workstation": ("工位电脑", WORKSTATION_COLUMNS, 14),
        "office": ("办公电脑", OFFICE_COLUMNS, 19),
    }

    ordered_records = sorted(records or [], key=_machine_sort_key) if sort_by_machine_name else (records or [])
    for record in ordered_records:
        kind = _record_type(record)
        if kind not in sheet_specs:
            continue
        sheet_name, columns, visible_cols = sheet_specs[kind]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # The provided workbook leaves the sequence column blank after the
        # first few rows.  It is purely positional, so complete it while
        # retaining any numbers that were already entered.
        for existing_row in range(2, ws.max_row + 1):
            if ws[f"B{existing_row}"].value is not None and ws[f"A{existing_row}"].value is None:
                ws[f"A{existing_row}"] = existing_row - 1
        machine_name = str(record.get("machine_name", "")).strip()
        if not machine_name:
            continue

        # Restrict scanning to the actual data columns; the template carries
        # formatting out to Excel's maximum column and therefore reports a
        # misleading max_column value.
        row_by_machine = {}
        for row in range(2, ws.max_row + 1):
            value = ws[f"B{row}"].value
            if value is not None and str(value).strip():
                row_by_machine[str(value).strip().upper()] = row
        row = row_by_machine.get(machine_name.upper())
        if row is None:
            row = ws.max_row + 1
            _copy_row_style(ws, max(2, row - 1), row, visible_cols)
            # Keep the template's sequence convention while tolerating gaps.
            numbers = [ws[f"A{r}"].value for r in range(2, row)
                       if isinstance(ws[f"A{r}"].value, (int, float))]
            ws[f"A{row}"] = int(max(numbers, default=0)) + 1

        for key, col in columns.items():
            value = record.get(key)
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            ws[f"{col}{row}"] = value.strip() if isinstance(value, str) else value

    wb.save(output_path)
    return output_path


WORKSTATION_HEADERS = [
    ('序号', None), ('计算机名', 'machine_name'), ('品牌', 'brand'),
    ('电脑型号', 'computer_model'), ('CPU', 'cpu'), ('网卡名', 'nic1_name'),
    ('物理地址', 'nic1_mac'), ('固态硬盘', 'disk_gb'), ('内存', 'memory_gb'),
    ('SN', 'sn_number'), ('系统类型', 'system_type'), ('操作系统', 'operating_system'),
    ('版本', 'version'), ('安装系统时间', 'install_time'),
]


def _machine_sort_key(record: dict):
    """Sort machine names naturally and place records without one last."""
    name = str(record.get('machine_name', '') or '').strip().upper()
    parts = re.split(r'(\d+)', name)
    natural = tuple((1, int(part)) if part.isdigit() else (0, part) for part in parts)
    return (not bool(name), natural, str(record.get('id', '')))
OFFICE_HEADERS = [
    ('序号', None), ('计算机名', 'machine_name'), ('品牌', 'brand'),
    ('电脑型号', 'computer_model'), ('CPU', 'cpu'), ('有线网卡名', 'nic1_name'),
    ('有线物理地址', 'nic1_mac'), ('无线网卡名', 'nic2_name'),
    ('无线物理地址', 'nic2_mac'), ('固态硬盘', 'disk_gb'), ('内存', 'memory_gb'),
    ('SN', 'sn_number'), ('部门', 'department'), ('姓名', 'person_name'),
    ('工号', 'employee_id'), ('安装系统时间', 'install_time'), ('系统类型', 'system_type'),
    ('操作系统', 'operating_system'), ('版本', 'version'),
]


def export_to_template(records: list, template_path: str = TEMPLATE_PATH,
                       output_path: str = None, embed_images: bool = True,
                       sort_by_machine_name: bool = True) -> str:
    """Create a fresh two-sheet workbook from historical OCR records.

    The source workbook is used only as a column/layout reference. It is never
    modified. Records are split by YCSPGW/YCSPBG machine prefix and duplicate
    serial numbers are emitted once.
    """
    if output_path is None:
        output_path = os.path.join(Config.DATA_FOLDER, 'YCSP电脑信息_已填写.xlsx')
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb = Workbook()
    first = wb.active
    first.title = '工位电脑'
    office = wb.create_sheet('办公电脑')

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, headers, rows):
        output_headers = list(headers) + [('原图', '_image')]
        for col, (title, key) in enumerate(output_headers, 1):
            cell = ws.cell(1, col, title)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            width = 34 if key == '_image' else max(12, min(34, len(title) * 2 + 8))
            ws.column_dimensions[get_column_letter(col)].width = width
        seen = set()
        out_row = 2
        ordered_rows = sorted(rows, key=_machine_sort_key) if sort_by_machine_name else rows
        for rec in ordered_rows:
            sn = str(rec.get('sn_number', '') or '').strip().upper()
            if sn and sn in seen:
                continue
            if sn: seen.add(sn)
            ws.row_dimensions[out_row].height = ROW_HEIGHT_PT
            for col, (_, key) in enumerate(output_headers, 1):
                if key == '_image':
                    cell = ws.cell(out_row, col)
                    cell.alignment = Alignment(vertical='center', horizontal='center')
                    cell.border = border
                    if embed_images:
                        thumb = _make_thumb(rec.get('image_path', ''))
                        if thumb:
                            try:
                                image = XLImage(thumb)
                                image.anchor = f'{get_column_letter(col)}{out_row}'
                                ws.add_image(image)
                            except Exception:
                                pass
                    continue
                value = out_row - 1 if key is None else rec.get(key, '')
                cell = ws.cell(out_row, col, value if value is not None else '')
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
            out_row += 1
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(output_headers))}{max(1, out_row - 1)}'

    workstation, office_rows = [], []
    for rec in records or []:
        kind = _record_type(rec)
        if kind == 'workstation': workstation.append(rec)
        elif kind == 'office': office_rows.append(rec)
    write_sheet(first, WORKSTATION_HEADERS, workstation)
    write_sheet(office, OFFICE_HEADERS, office_rows)
    wb.save(output_path)
    return output_path
